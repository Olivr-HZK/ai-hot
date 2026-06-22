from __future__ import annotations

import argparse
import json
import random
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(SCRIPT_DIR))

from env_utils import load_env
from feedback_rules import is_excluded_by_rules, load_feedback_rules, passes_include_keywords, video_haystack
from phase1_scrape_x import normalize_x_hotspot, parse_x_datetime


DEFAULT_SOURCE_DIR = BASE_DIR / "x_history_eval" / "sources"
DEFAULT_RAW_SOURCE_DIR = BASE_DIR.parent / "X-rader-AInews" / "trend-scrap" / "x-scraper" / "data" / "raw"
DEFAULT_OUTPUT = BASE_DIR / "x_history_eval" / "x_history_scored.xlsx"
DEFAULT_JSON_OUTPUT = BASE_DIR / "x_history_eval" / "x_history_scored.json"
DEFAULT_RAW_SAMPLE_OUTPUT = BASE_DIR / "x_history_eval" / "x_raw_unfiltered_sample_5pct.json"


def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return default


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def post_key(post: dict[str, Any]) -> str:
    return clean_text(post.get("url")) or f"x:{post.get('id')}"


def first_present(item: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return value
    return ""


def normalize_history_post(raw: dict[str, Any], source_path: Path) -> dict[str, Any] | None:
    url = first_present(raw, ["url", "post_url", "hotspotUrl"])
    post_id = first_present(raw, ["id", "post_id", "tweet_id"])
    if not url and not post_id:
        return None
    author = raw.get("author")
    if isinstance(author, str):
        author_obj = {"username": author, "display_name": author}
    elif isinstance(author, dict):
        author_obj = author
    else:
        author_obj = {"username": "", "display_name": ""}
    text = first_present(raw, ["text", "post_title", "title", "trend_title"])
    summary = first_present(raw, ["post_summary", "summary", "background", "hotspotIntro"])
    normalized = {
        "id": str(post_id or ""),
        "url": str(url or ""),
        "text": str(text or ""),
        "author": author_obj,
        "created_at": first_present(raw, ["created_at", "createTime", "createTimeISO"]),
        "view_count": safe_int(first_present(raw, ["view_count", "views", "playCount"])),
        "like_count": safe_int(first_present(raw, ["like_count", "likes", "diggCount", "likeCount"])),
        "reply_count": safe_int(first_present(raw, ["reply_count", "replies", "commentCount"])),
        "retweet_count": safe_int(first_present(raw, ["retweet_count", "retweets", "retweetCount"])),
        "video_duration_seconds": safe_int(raw.get("video_duration_seconds")),
        "post_summary": str(summary or ""),
        "publishDays": safe_int(first_present(raw, ["publishDays", "days_ago"]), 0),
        "stage2_target_category": str(first_present(raw, ["stage2_target_category", "category_tag"]) or ""),
        "history_source_file": str(source_path),
        "history_sample_type": str(raw.get("history_sample_type") or "已筛热点"),
    }
    if not normalized["text"] and normalized["post_summary"]:
        normalized["text"] = normalized["post_summary"]
    return normalized


def extract_posts_from_report(data: dict[str, Any], source_path: Path) -> list[dict[str, Any]]:
    posts: list[dict[str, Any]] = []
    for item in data.get("items", []) if isinstance(data.get("items"), list) else []:
        if isinstance(item, dict) and (item.get("post_url") or item.get("post_id")):
            posts.append({**item, "history_report_title": data.get("title", "")})
        for nested_key in ["representative_posts", "reference_links", "links"]:
            nested = item.get(nested_key) if isinstance(item, dict) else None
            if not isinstance(nested, list):
                continue
            for nested_item in nested:
                if isinstance(nested_item, dict):
                    posts.append(
                        {
                            **nested_item,
                            "category_tag": item.get("category_tag", ""),
                            "trend_title": item.get("trend_title", ""),
                            "background": nested_item.get("post_summary") or item.get("background", ""),
                            "history_report_title": data.get("title", ""),
                        }
                    )
    return [post for post in posts if normalize_history_post(post, source_path)]


def load_history_posts(source_dir: Path) -> list[dict[str, Any]]:
    posts: list[dict[str, Any]] = []
    for path in sorted(source_dir.rglob("*.json")):
        if "raw" in path.parts:
            continue
        if path.name not in {"filtered-result.json", "report_per_post.json", "report.json"}:
            continue
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        candidates: list[dict[str, Any]] = []
        if isinstance(data, list):
            candidates = [item for item in data if isinstance(item, dict)]
        elif isinstance(data, dict):
            candidates = extract_posts_from_report(data, path)
        for candidate in candidates:
            post = normalize_history_post(candidate, path)
            if post:
                posts.append(post)
    return posts


def load_raw_unfiltered_posts(raw_source_dir: Path, excluded_keys: set[str]) -> list[dict[str, Any]]:
    posts: list[dict[str, Any]] = []
    if not raw_source_dir.exists():
        return posts
    for path in sorted(raw_source_dir.glob("*.json")):
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        pages = data if isinstance(data, list) else [data]
        for page in pages:
            if not isinstance(page, dict):
                continue
            filtered_ids = {str(item) for item in page.get("filtered_tweet_ids", [])}
            tweets = page.get("tweets") if isinstance(page.get("tweets"), list) else []
            for tweet in tweets:
                if not isinstance(tweet, dict):
                    continue
                if str(tweet.get("id") or "") in filtered_ids:
                    continue
                post = normalize_history_post(
                    {
                        **tweet,
                        "history_sample_type": "raw未初筛抽样",
                        "search_term": tweet.get("search_term") or page.get("search_term") or "",
                    },
                    path,
                )
                if not post:
                    continue
                key = post_key(post)
                if not key or key in excluded_keys:
                    continue
                posts.append(post)
    return posts


def dedupe_posts(posts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[str, dict[str, Any]] = {}
    for post in posts:
        key = post_key(post)
        if not key:
            continue
        existing = deduped.get(key)
        if not existing:
            deduped[key] = post
            continue
        if len(clean_text(post.get("post_summary"))) > len(clean_text(existing.get("post_summary"))):
            deduped[key] = post
    return list(deduped.values())


def keyword_hits(haystack: str, keywords: list[str]) -> list[str]:
    return [keyword for keyword in keywords if keyword.strip() and keyword.strip().lower() in haystack]


def evaluate_post(post: dict[str, Any], rules: dict[str, Any]) -> dict[str, Any]:
    hotspot = normalize_x_hotspot(post, rules=rules)
    filters = rules.get("filters", {})
    haystack = video_haystack(hotspot, include_summary=True)
    include_hits = keyword_hits(haystack, filters.get("include_keywords", []))
    exclude_hits = keyword_hits(haystack, filters.get("exclude_keywords", []))
    preferred_hits = keyword_hits(haystack, filters.get("preferred_keywords", []))
    deprioritized_hits = keyword_hits(haystack, filters.get("deprioritized_keywords", []))
    group_hits = []
    for group in filters.get("exclude_keyword_groups", []):
        cleaned = [str(item).strip().lower() for item in group if str(item).strip()]
        if cleaned and all(item in haystack for item in cleaned):
            group_hits.append("+".join(cleaned))
    accepted = passes_include_keywords(hotspot, rules) and not is_excluded_by_rules(hotspot, rules, include_summary=True)
    reasons = []
    if not include_hits and filters.get("include_keywords"):
        reasons.append("未命中包含关键词")
    if exclude_hits:
        reasons.append("命中排除关键词: " + ", ".join(exclude_hits))
    if group_hits:
        reasons.append("命中组合排除: " + ", ".join(group_hits))
    if accepted and preferred_hits:
        reasons.append("命中加权关键词: " + ", ".join(preferred_hits))
    if accepted and deprioritized_hits:
        reasons.append("命中降权关键词: " + ", ".join(deprioritized_hits))
    if not reasons:
        reasons.append("通过当前筛选规则")
    dt = parse_x_datetime(hotspot)
    author = hotspot.get("authorMeta") or {}
    return {
        "历史来源文件": str(post.get("history_source_file") or ""),
        "样本类型": str(post.get("history_sample_type") or "已筛热点"),
        "帖子ID": str(hotspot.get("id") or ""),
        "发布时间": dt.isoformat(sep=" ") if dt else "",
        "作者": author.get("nickName") or author.get("name") or "",
        "热点链接": hotspot.get("hotspotUrl") or "",
        "正文": hotspot.get("text") or "",
        "历史摘要": hotspot.get("video_summary") or "",
        "浏览量": hotspot.get("playCount") or 0,
        "点赞数": hotspot.get("diggCount") or 0,
        "回复数": hotspot.get("commentCount") or 0,
        "转发数": hotspot.get("retweetCount") or 0,
        "发布天数": hotspot.get("publishDays") or 0,
        "热度评分": hotspot.get("heatValue") or 0,
        "筛选判断": "接受" if accepted else "否决",
        "判断原因": "；".join(reasons),
        "命中加权关键词": ", ".join(preferred_hits),
        "命中降权关键词": ", ".join(deprioritized_hits),
        "原始历史分类": hotspot.get("stage2_target_category") or "",
    }


def write_excel(rows: list[dict[str, Any]], output: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install dependencies from requirements.txt") from exc
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "x_history_scored"
    headers = list(rows[0].keys()) if rows else [
        "历史来源文件",
        "样本类型",
        "帖子ID",
        "发布时间",
        "作者",
        "热点链接",
        "正文",
        "历史摘要",
        "浏览量",
        "点赞数",
        "回复数",
        "转发数",
        "发布天数",
        "热度评分",
        "筛选判断",
        "判断原因",
        "命中加权关键词",
        "命中降权关键词",
        "原始历史分类",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for index, header in enumerate(headers, 1):
        width = 18
        if header in {"正文", "历史摘要", "判断原因"}:
            width = 48
        elif header in {"热点链接", "历史来源文件"}:
            width = 36
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate historical X hotspots with current scoring and filters")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--raw-source-dir", type=Path, default=DEFAULT_RAW_SOURCE_DIR)
    parser.add_argument("--raw-sample-percent", type=float, default=0.05)
    parser.add_argument("--random-seed", type=int, default=20260513)
    parser.add_argument("--raw-sample-output", type=Path, default=DEFAULT_RAW_SAMPLE_OUTPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--json-output", type=Path, default=DEFAULT_JSON_OUTPUT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    load_env()
    rules = load_feedback_rules()
    history_posts = dedupe_posts(load_history_posts(args.source_dir))
    history_keys = {post_key(post) for post in history_posts if post_key(post)}
    raw_pool = load_raw_unfiltered_posts(args.raw_source_dir, history_keys)
    sample_size = round(len(raw_pool) * max(0.0, args.raw_sample_percent))
    raw_sample = random.Random(args.random_seed).sample(raw_pool, min(sample_size, len(raw_pool))) if raw_pool else []
    args.raw_sample_output.parent.mkdir(parents=True, exist_ok=True)
    args.raw_sample_output.write_text(json.dumps(raw_sample, ensure_ascii=False, indent=2), encoding="utf-8")
    posts = history_posts + raw_sample
    rows = [evaluate_post(post, rules) for post in posts]
    rows.sort(key=lambda row: float(row.get("热度评分") or 0), reverse=True)
    args.json_output.parent.mkdir(parents=True, exist_ok=True)
    args.json_output.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(rows, args.output)
    accepted = len([row for row in rows if row.get("筛选判断") == "接受"])
    rejected = len(rows) - accepted
    print(
        json.dumps(
            {
                "source_dir": str(args.source_dir),
                "raw_source_dir": str(args.raw_source_dir),
                "screened_hotspots": len(history_posts),
                "raw_unfiltered_pool": len(raw_pool),
                "raw_sample_percent": args.raw_sample_percent,
                "raw_sample_size": len(raw_sample),
                "random_seed": args.random_seed,
                "evaluated_rows": len(rows),
                "accepted": accepted,
                "rejected": rejected,
                "raw_sample_output": str(args.raw_sample_output),
                "json_output": str(args.json_output),
                "excel_output": str(args.output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
